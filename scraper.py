"""Collect job seeker leads from public forums and export them to Excel."""

from __future__ import annotations

import html
import logging
import os
import random
import re
import time
import unicodedata
from datetime import UTC, datetime, timedelta
from typing import Any
from urllib.parse import urljoin

import pandas as pd
import praw
import requests
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "job_seekers.xlsx"
USER_AGENT = "job-seeker-scraper/1.0"
REQUEST_TIMEOUT = 30
SLEEP_RANGE_SECONDS = (1.0, 2.0)

REDDIT_SUBREDDITS = [
    "foicontratado",
    "brdev",
    "devbrasil",
    "programacao",
    "cscareerquestions",
]

REDDIT_SEARCH_QUERIES = [
    "procurando emprego",
    "procuro emprego",
    "buscando vaga",
    "arrumar emprego",
    "recolocacao",
    "desempregado",
    "fui demitido",
    "open to work",
    "looking for a job",
    "looking for work",
    "need a job",
    "laid off",
    "available for hire",
]

HN_SEARCH_QUERIES = [
    "open to work",
    "looking for a job",
    "looking for work",
    "available for hire",
    "seeking opportunities",
    "need a job",
    "procurando emprego",
    "buscando vaga",
]

SEEKER_KEYWORDS_PT = [
    "procurando emprego",
    "procuro emprego",
    "procurando trabalho",
    "procuro trabalho",
    "buscando emprego",
    "busco emprego",
    "buscando vaga",
    "busco vaga",
    "procuro vaga",
    "em busca de emprego",
    "em busca de uma vaga",
    "estou procurando emprego",
    "estou procurando trabalho",
    "estou em busca de uma vaga",
    "to procurando emprego",
    "to procurando vaga",
    "arrumar emprego",
    "achar emprego",
    "conseguir emprego",
    "conseguir uma vaga",
    "buscar trabalho",
    "buscando recolocacao",
    "recolocacao",
    "realocacao",
    "disponivel para trabalhar",
    "disponivel para freelas",
    "disponivel para trabalho",
    "aberto a oportunidades",
    "aberta a oportunidades",
    "preciso de emprego",
    "preciso trabalhar",
    "desempregado",
    "sem emprego",
    "fui demitido",
    "fui demitida",
]

SEEKER_KEYWORDS_EN = [
    "open to work",
    "looking for a job",
    "looking for work",
    "looking for employment",
    "looking for opportunities",
    "searching for a job",
    "searching for work",
    "seeking opportunities",
    "seeking a job",
    "seeking a new role",
    "seeking new role",
    "seeking a position",
    "available for work",
    "available for hire",
    "actively looking",
    "actively job hunting",
    "job hunting",
    "need a job",
    "need work",
    "unemployed",
    "between jobs",
    "laid off",
    "laid off and looking",
]

RECRUITER_SIGNALS = [
    "we are hiring",
    "we're hiring",
    "hiring now",
    "hiring for",
    "vaga aberta",
    "vagas abertas",
    "estamos contratando",
    "estou contratando",
    "contratando",
    "recruiter here",
    "recruiter",
    "talent acquisition",
    "send your resume",
    "send us your resume",
    "job opening",
    "job openings",
]

HARD_RECRUITER_SIGNALS = {
    "we are hiring",
    "we're hiring",
    "hiring now",
    "hiring for",
    "vaga aberta",
    "vagas abertas",
    "estamos contratando",
    "estou contratando",
}

ALL_SEEKER_KEYWORDS = SEEKER_KEYWORDS_PT + SEEKER_KEYWORDS_EN
SEEKER_REGEX_PATTERNS = [
    (
        "job-search intent pattern",
        re.compile(r"\b(arrumar|achar|conseguir|buscar|procurar)\b.{0,40}\b(emprego|vaga|trampo|trabalho)\b"),
    ),
    (
        "unemployment pattern",
        re.compile(r"\b(desempregado|sem emprego|fui demitid[oa]|laid off|unemployed|between jobs)\b"),
    ),
]
OUTPUT_COLUMNS = [
    "source",
    "forum",
    "author",
    "title",
    "content",
    "url",
    "date",
    "location_country",
    "technologies",
    "experience",
    "keywords_matched",
]
SOURCE_ORDER = ["Reddit", "HackerNews", "TabNews", "GUJ"]
SOURCE_SHEET_NAMES = {
    "Reddit": "Reddit",
    "HackerNews": "HackerNews",
    "TabNews": "TabNews",
    "GUJ": "GUJ",
}

TABNEWS_BASE_URL = "https://www.tabnews.com.br/api/v1"
GUJ_CATEGORY_URL = "https://www.guj.com.br/c/empregos"
GUJ_FALLBACK_URL = "https://www.guj.com.br/"
GUJ_MAX_TOPIC_LINKS = 20
TABNEWS_MAX_PAGES = 5
TABNEWS_PAGE_SIZE = 100
HN_MAX_PAGES_PER_QUERY = 2
HN_MONTHLY_THREADS_TO_SCAN = 6
HN_RECENT_LOOKBACK_DAYS = 45

HN_FIELD_NAME_MAP = {
    "location": "location",
    "technologies": "technologies",
    "technology": "technologies",
    "stack": "technologies",
    "looking for": "looking_for",
    "resume/cv": "resume",
    "resume": "resume",
    "cv": "resume",
    "email": "email",
    "github": "github",
    "website": "website",
    "portfolio": "website",
    "linkedin": "linkedin",
}

HN_STRUCTURED_LABEL_PATTERN = re.compile(
    r"(?P<label>Location|Remote|Willing to relocate|Technologies|Technology|Stack|Looking for|R\S*sum\S*\s*/\s*CV|Résumé\s*/\s*CV|Resume\s*/\s*CV|Résumé/CV|Resume/CV|Resume|CV|Email|Github|GitHub|Website|Portfolio|LinkedIn)\s*:\s*",
    flags=re.IGNORECASE,
)

COUNTRY_ALIAS_MAP = {
    "Australia": ["australia", "sydney", "melbourne", "brisbane", "perth"],
    "Austria": ["austria", "vienna"],
    "Belgium": ["belgium", "brussels"],
    "Bulgaria": ["bulgaria", "sofia"],
    "Brazil": ["brazil", "brasil", "sao paulo", "rio de janeiro", "belo horizonte", "curitiba", "porto alegre"],
    "Canada": ["canada", "toronto", "vancouver", "montreal", "ottawa", "calgary"],
    "Denmark": ["denmark", "copenhagen"],
    "Finland": ["finland", "helsinki"],
    "France": ["france", "paris", "lyon", "marseille"],
    "Germany": ["germany", "berlin", "munich", "hamburg", "frankfurt", "cologne"],
    "India": ["india", "bangalore", "bengaluru", "mumbai", "delhi", "new delhi", "hyderabad", "pune", "chennai"],
    "Ireland": ["ireland", "dublin", "cork"],
    "Japan": ["japan", "tokyo", "osaka", "kyoto"],
    "Mexico": ["mexico", "mexico city", "guadalajara", "monterrey"],
    "Netherlands": ["netherlands", "amsterdam", "rotterdam", "utrecht", "the hague"],
    "New Zealand": ["new zealand", "auckland", "wellington"],
    "Norway": ["norway", "oslo"],
    "Poland": ["poland", "warsaw", "krakow", "wroclaw"],
    "Portugal": ["portugal", "lisbon", "porto"],
    "Singapore": ["singapore"],
    "Spain": ["spain", "madrid", "barcelona", "valencia"],
    "Sweden": ["sweden", "stockholm", "gothenburg"],
    "Switzerland": ["switzerland", "zurich", "geneva"],
    "United Kingdom": ["united kingdom", "uk", "england", "london", "manchester", "edinburgh", "glasgow", "birmingham"],
    "United States": ["united states", "usa", "san francisco", "new york", "nyc", "seattle", "austin", "boston", "chicago", "los angeles", "bay area", "atlanta", "miami", "washington dc"],
}

EXPERIENCE_ROLE_PATTERNS = [
    ("Product Engineer", ["product engineer"]),
    ("Data Engineer", ["data engineer", "analytics engineer", "etl engineer", "data platform engineer"]),
    ("Data Scientist", ["data scientist", "applied scientist", "research scientist"]),
    ("ML Engineer", ["machine learning engineer", "ml engineer", "ai engineer", "llm engineer", "applied ai engineer"]),
    ("AI Researcher", ["ai researcher", "research engineer"]),
    ("Backend Engineer", ["backend engineer", "backend developer", "backend software engineer"]),
    ("Frontend Engineer", ["frontend engineer", "front-end engineer", "frontend developer", "front end developer", "frontend software engineer"]),
    ("Full-Stack Engineer", ["full-stack engineer", "full stack engineer", "full-stack developer", "full stack developer", "full-stack", "full stack"]),
    ("Software Developer", ["software developer", "developer", "software engineer", "swe"]),
    ("Mobile Engineer", ["mobile engineer", "ios engineer", "android engineer"]),
    ("DevOps / Platform Engineer", ["devops engineer", "platform engineer", "site reliability engineer", "sre", "cloud engineer", "infra engineer"]),
    ("Quant / Finance", ["quant", "quantitative", "trading", "finance"]),
]


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )


def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    return session


def rate_limit_sleep() -> None:
    time.sleep(random.uniform(*SLEEP_RANGE_SECONDS))


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    decomposed = unicodedata.normalize("NFKD", value.casefold())
    without_accents = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", without_accents).strip()


def truncate_text(value: str | None, max_length: int = 600) -> str:
    if not value:
        return ""
    cleaned = re.sub(r"\s+", " ", value).strip()
    if len(cleaned) <= max_length:
        return cleaned
    return cleaned[: max_length - 3].rstrip() + "..."


def sanitize_excel_text(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", value)


def html_to_text(value: str | None, separator: str = " ") -> str:
    if not value:
        return ""
    soup = BeautifulSoup(html.unescape(value), "html.parser")
    return soup.get_text(separator, strip=True)


def html_to_lines(value: str | None) -> list[str]:
    if not value:
        return []
    raw_text = html_to_text(value, separator="\n")
    lines: list[str] = []
    for raw_line in raw_text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        if line:
            lines.append(line)
    return lines


def find_matches(text: str | None, candidates: list[str]) -> list[str]:
    normalized_text = normalize_text(text)
    matches: list[str] = []
    for phrase in candidates:
        if normalize_text(phrase) in normalized_text:
            matches.append(phrase)
    return matches


def find_regex_matches(text: str | None) -> list[str]:
    normalized_text = normalize_text(text)
    matches: list[str] = []
    for label, pattern in SEEKER_REGEX_PATTERNS:
        if pattern.search(normalized_text):
            matches.append(label)
    return matches


def parse_iso_date(value: str | None) -> str:
    if not value:
        return ""
    normalized = value.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized).date().isoformat()
    except ValueError:
        return ""


def parse_iso_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    normalized = value.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=UTC)
    return parsed.astimezone(UTC)


def parse_timestamp_date(timestamp: float | int | None) -> str:
    if timestamp is None:
        return ""
    return datetime.fromtimestamp(timestamp, tz=UTC).date().isoformat()


def parse_guj_date(value: str | None) -> str:
    if not value:
        return ""

    month_map = {
        "jan": 1,
        "janeiro": 1,
        "fev": 2,
        "fevereiro": 2,
        "mar": 3,
        "marco": 3,
        "abril": 4,
        "abr": 4,
        "maio": 5,
        "jun": 6,
        "junho": 6,
        "jul": 7,
        "julho": 7,
        "ago": 8,
        "agosto": 8,
        "set": 9,
        "setembro": 9,
        "out": 10,
        "outubro": 10,
        "nov": 11,
        "novembro": 11,
        "dez": 12,
        "dezembro": 12,
    }

    normalized = normalize_text(value).replace(".", "")
    match = re.search(r"(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})", normalized)
    if not match:
        return ""
    day, month_name, year = match.groups()
    month = month_map.get(month_name)
    if not month:
        return ""
    try:
        return datetime(int(year), month, int(day)).date().isoformat()
    except ValueError:
        return ""


def select_text(*values: str | None) -> str:
    for value in values:
        if value and str(value).strip():
            return str(value).strip()
    return ""


def clean_structured_value(value: str) -> str:
    value = re.sub(r"\s+", " ", value).strip()
    value = re.sub(r"\s*,\s*", ", ", value)
    return value.strip(" ;,")


def extract_hn_structured_fields(content: str | None) -> dict[str, str]:
    fields: dict[str, str] = {}
    searchable_text = "\n".join(html_to_lines(content))
    matches = list(HN_STRUCTURED_LABEL_PATTERN.finditer(searchable_text))

    for index, match in enumerate(matches):
        raw_label = match.group("label")
        normalized_label = normalize_text(raw_label)
        field_name = HN_FIELD_NAME_MAP.get(normalized_label)
        if not field_name:
            continue

        value_start = match.end()
        value_end = matches[index + 1].start() if index + 1 < len(matches) else len(searchable_text)
        raw_value = searchable_text[value_start:value_end]
        cleaned_value = clean_structured_value(raw_value)
        if cleaned_value:
            fields[field_name] = cleaned_value

    return fields


def infer_location_country(location_value: str) -> str:
    normalized_location = normalize_text(location_value)
    if not normalized_location:
        return ""

    countries: list[str] = []
    for country, aliases in COUNTRY_ALIAS_MAP.items():
        for alias in aliases:
            normalized_alias = normalize_text(alias)
            if re.search(rf"\b{re.escape(normalized_alias)}\b", normalized_location):
                countries.append(country)
                break

    if re.search(r",\s*[A-Z]{2}\b", location_value) and "United States" not in countries:
        countries.append("United States")

    if not countries:
        fragments = [clean_structured_value(fragment) for fragment in location_value.split(",") if clean_structured_value(fragment)]
        if len(fragments) >= 2:
            last_fragment = fragments[-1]
            if re.fullmatch(r"[A-Za-z .'-]{3,}", last_fragment):
                countries.append(last_fragment)

    return ", ".join(dict.fromkeys(countries))


def infer_experience_profile(*text_parts: str) -> str:
    searchable_text = normalize_text(" ".join(part for part in text_parts if part))
    if not searchable_text:
        return ""

    matches: list[str] = []
    for role_name, patterns in EXPERIENCE_ROLE_PATTERNS:
        for pattern in patterns:
            normalized_pattern = normalize_text(pattern)
            if re.search(rf"\b{re.escape(normalized_pattern)}\b", searchable_text):
                matches.append(role_name)
                break

    return ", ".join(dict.fromkeys(matches))


def extract_hn_candidate_metadata(title: str, raw_content: str | None, plain_content: str) -> dict[str, str]:
    structured_fields = extract_hn_structured_fields(raw_content or plain_content)
    location_country = infer_location_country(structured_fields.get("location", ""))
    technologies = structured_fields.get("technologies", "")
    experience = infer_experience_profile(
        title,
        plain_content,
        structured_fields.get("looking_for", ""),
        technologies,
    )

    return {
        "location_country": location_country,
        "technologies": technologies,
        "experience": experience,
    }


def should_include_item(
    title: str,
    content: str,
    context_keywords: list[str] | None = None,
    ignore_soft_recruiter_signals: bool = False,
) -> tuple[bool, list[str], list[str]]:
    full_text = " ".join(part for part in [title, content] if part).strip()
    seeker_matches = find_matches(full_text, ALL_SEEKER_KEYWORDS)
    regex_matches = find_regex_matches(full_text)
    recruiter_matches = find_matches(full_text, RECRUITER_SIGNALS)
    combined_keywords = seeker_matches[:]

    for regex_match in regex_matches:
        if regex_match not in combined_keywords:
            combined_keywords.append(regex_match)

    if context_keywords:
        for keyword in context_keywords:
            if keyword not in combined_keywords:
                combined_keywords.append(keyword)

    if not combined_keywords:
        return False, [], recruiter_matches

    if recruiter_matches:
        hard_matches = [signal for signal in recruiter_matches if signal in HARD_RECRUITER_SIGNALS]
        if hard_matches and not seeker_matches and not regex_matches:
            return False, combined_keywords, recruiter_matches
        if recruiter_matches and not seeker_matches and not regex_matches and not ignore_soft_recruiter_signals:
            return False, combined_keywords, recruiter_matches

    return True, combined_keywords, recruiter_matches


def build_record(
    *,
    source: str,
    forum: str,
    author: str,
    title: str,
    content: str,
    url: str,
    date: str,
    keywords_matched: list[str],
    location_country: str = "",
    technologies: str = "",
    experience: str = "",
) -> dict[str, str]:
    return {
        "source": sanitize_excel_text(source),
        "forum": sanitize_excel_text(forum),
        "author": sanitize_excel_text(author.strip() if author else ""),
        "title": sanitize_excel_text(title.strip() if title else ""),
        "content": sanitize_excel_text(truncate_text(content)),
        "url": sanitize_excel_text(url.strip()),
        "date": sanitize_excel_text(date),
        "location_country": sanitize_excel_text(location_country.strip()),
        "technologies": sanitize_excel_text(technologies.strip()),
        "experience": sanitize_excel_text(experience.strip()),
        "keywords_matched": sanitize_excel_text(", ".join(dict.fromkeys(keywords_matched))),
    }


def deduplicate_records(records: list[dict[str, str]]) -> list[dict[str, str]]:
    seen_urls: set[str] = set()
    unique_records: list[dict[str, str]] = []
    for record in records:
        url = record.get("url", "").strip()
        if not url or url in seen_urls:
            continue
        seen_urls.add(url)
        unique_records.append(record)
    return unique_records


def make_dataframe(records: list[dict[str, str]]) -> pd.DataFrame:
    dataframe = pd.DataFrame(records, columns=OUTPUT_COLUMNS)
    if dataframe.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    return dataframe.sort_values("date", ascending=False, na_position="last").reset_index(drop=True)


def auto_adjust_column_widths(writer: pd.ExcelWriter) -> None:
    for worksheet in writer.book.worksheets:
        for idx, column_cells in enumerate(worksheet.iter_cols(1, worksheet.max_column), start=1):
            max_length = 0
            for cell in column_cells:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            adjusted = min(max_length + 2, 80)
            worksheet.column_dimensions[get_column_letter(idx)].width = max(adjusted, 12)


def export_to_excel(records_by_source: dict[str, list[dict[str, str]]], output_path: str = OUTPUT_FILE) -> None:
    consolidated_records: list[dict[str, str]] = []
    for source_name in SOURCE_ORDER:
        consolidated_records.extend(records_by_source.get(source_name, []))

    consolidated_records = deduplicate_records(consolidated_records)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        make_dataframe(consolidated_records).to_excel(writer, sheet_name="Leads", index=False)

        for source_name in SOURCE_ORDER:
            source_records = deduplicate_records(records_by_source.get(source_name, []))
            make_dataframe(source_records).to_excel(
                writer,
                sheet_name=SOURCE_SHEET_NAMES[source_name],
                index=False,
            )

        auto_adjust_column_widths(writer)


def request_json(
    session: requests.Session,
    url: str,
    *,
    params: dict[str, Any] | None = None,
) -> Any:
    response = session.get(url, params=params, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    rate_limit_sleep()
    return response.json()


def request_html(
    session: requests.Session,
    url: str,
    *,
    params: dict[str, Any] | None = None,
) -> BeautifulSoup:
    response = session.get(url, params=params, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    rate_limit_sleep()
    return BeautifulSoup(response.text, "html.parser")


def fetch_reddit_leads() -> list[dict[str, str]]:
    client_id = os.getenv("REDDIT_CLIENT_ID")
    client_secret = os.getenv("REDDIT_CLIENT_SECRET")
    if not client_id or not client_secret:
        logging.warning(
            "Reddit credentials not found. Set REDDIT_CLIENT_ID and REDDIT_CLIENT_SECRET to enable Reddit collection."
        )
        return []

    reddit = praw.Reddit(
        client_id=client_id,
        client_secret=client_secret,
        user_agent=USER_AGENT,
    )
    reddit.read_only = True

    records: list[dict[str, str]] = []

    for subreddit_name in REDDIT_SUBREDDITS:
        logging.info("Scanning Reddit subreddit: %s", subreddit_name)
        subreddit = reddit.subreddit(subreddit_name)
        subreddit_matches = 0

        for submission in subreddit.new(limit=200):
            title = submission.title or ""
            content = submission.selftext or ""
            include, keywords_matched, _ = should_include_item(title, content)
            if not include:
                continue

            records.append(
                build_record(
                    source="Reddit",
                    forum=subreddit_name,
                    author=str(submission.author or ""),
                    title=title,
                    content=content or submission.url,
                    url=f"https://www.reddit.com{submission.permalink}",
                    date=parse_timestamp_date(submission.created_utc),
                    keywords_matched=keywords_matched,
                )
            )
            subreddit_matches += 1

        rate_limit_sleep()

        for query in REDDIT_SEARCH_QUERIES:
            logging.info("Searching Reddit r/%s for query: %s", subreddit_name, query)
            try:
                submissions = subreddit.search(query, sort="new", limit=50)
                for submission in submissions:
                    title = submission.title or ""
                    content = submission.selftext or ""
                    include, keywords_matched, _ = should_include_item(title, content)
                    if not include:
                        continue

                    records.append(
                        build_record(
                            source="Reddit",
                            forum=subreddit_name,
                            author=str(submission.author or ""),
                            title=title,
                            content=content or submission.url,
                            url=f"https://www.reddit.com{submission.permalink}",
                            date=parse_timestamp_date(submission.created_utc),
                            keywords_matched=keywords_matched,
                        )
                    )
                    subreddit_matches += 1
            except Exception:
                logging.exception(
                    "Reddit search failed for r/%s with query %r",
                    subreddit_name,
                    query,
                )
            rate_limit_sleep()

        logging.info("Finished Reddit subreddit: %s (%s matched leads)", subreddit_name, subreddit_matches)

    return deduplicate_records(records)


def fetch_hackernews_leads() -> list[dict[str, str]]:
    session = build_session()
    records: list[dict[str, str]] = []
    recent_cutoff = datetime.now(UTC) - timedelta(days=HN_RECENT_LOOKBACK_DAYS)

    logging.info(
        "Scanning Hacker News keyword matches via Algolia API (recent window: last %s days)",
        HN_RECENT_LOOKBACK_DAYS,
    )
    for query in HN_SEARCH_QUERIES:
        for page in range(HN_MAX_PAGES_PER_QUERY):
            payload = request_json(
                session,
                "https://hn.algolia.com/api/v1/search_by_date",
                params={
                    "query": query,
                    "tags": "story,comment",
                    "page": page,
                    "hitsPerPage": 100,
                },
            )

            hits = payload.get("hits", [])
            if not hits:
                break

            hit_timestamps = [parse_iso_datetime(hit.get("created_at")) for hit in hits]
            for hit in hits:
                created_at = parse_iso_datetime(hit.get("created_at"))
                if created_at and created_at < recent_cutoff:
                    continue

                title = select_text(hit.get("title"), hit.get("story_title"), "[HN Comment]")
                raw_content = select_text(hit.get("comment_text"), hit.get("story_text"))
                content = html_to_text(raw_content)
                include, keywords_matched, _ = should_include_item(title, content)
                if not include:
                    continue

                item_id = hit.get("objectID") or hit.get("story_id")
                fallback_url = f"https://news.ycombinator.com/item?id={item_id}" if item_id else ""
                url = select_text(hit.get("url"), hit.get("story_url"), fallback_url)
                metadata = extract_hn_candidate_metadata(title, raw_content, content)

                records.append(
                    build_record(
                        source="HackerNews",
                        forum="HN Search",
                        author=str(hit.get("author") or ""),
                        title=title,
                        content=content,
                        url=url,
                        date=parse_iso_date(hit.get("created_at")),
                        keywords_matched=keywords_matched,
                        location_country=metadata["location_country"],
                        technologies=metadata["technologies"],
                        experience=metadata["experience"],
                    )
                )

            oldest_hit = min((timestamp for timestamp in hit_timestamps if timestamp), default=None)
            if oldest_hit and oldest_hit < recent_cutoff:
                break

    logging.info('Scanning Hacker News "Who wants to be hired?" monthly threads')
    thread_payload = request_json(
        session,
        "https://hn.algolia.com/api/v1/search_by_date",
        params={
            "query": "\"Who wants to be hired\"",
            "tags": "story",
            "hitsPerPage": 20,
        },
    )

    thread_hits = [
        hit
        for hit in thread_payload.get("hits", [])
        if "who wants to be hired" in normalize_text(hit.get("title", ""))
    ][:HN_MONTHLY_THREADS_TO_SCAN]

    for thread in thread_hits:
        story_id = thread.get("objectID")
        if not story_id:
            continue

        comments_payload = request_json(
            session,
            "https://hn.algolia.com/api/v1/search_by_date",
            params={
                "tags": f"comment,story_{story_id}",
                "hitsPerPage": 1000,
            },
        )

        for comment in comments_payload.get("hits", []):
            if str(comment.get("parent_id")) != str(comment.get("story_id")):
                continue
            created_at = parse_iso_datetime(comment.get("created_at"))
            if created_at and created_at < recent_cutoff:
                continue

            raw_content = comment.get("comment_text")
            content = html_to_text(raw_content)
            title = thread.get("title") or "Ask HN: Who wants to be hired?"
            include, keywords_matched, recruiter_matches = should_include_item(
                title,
                content,
                context_keywords=["who wants to be hired thread"],
                ignore_soft_recruiter_signals=True,
            )
            if not include:
                continue
            if any(match in HARD_RECRUITER_SIGNALS for match in recruiter_matches):
                continue

            metadata = extract_hn_candidate_metadata(title, raw_content, content)

            records.append(
                build_record(
                    source="HackerNews",
                    forum="Who wants to be hired?",
                    author=str(comment.get("author") or ""),
                    title=title,
                    content=content,
                    url=f"https://news.ycombinator.com/item?id={comment.get('objectID')}",
                    date=parse_iso_date(comment.get("created_at")),
                    keywords_matched=keywords_matched,
                    location_country=metadata["location_country"],
                    technologies=metadata["technologies"],
                    experience=metadata["experience"],
                )
            )

    return deduplicate_records(records)


def fetch_tabnews_leads() -> list[dict[str, str]]:
    session = build_session()
    records: list[dict[str, str]] = []

    logging.info("Scanning TabNews recent contents")
    for page in range(1, TABNEWS_MAX_PAGES + 1):
        items = request_json(
            session,
            f"{TABNEWS_BASE_URL}/contents",
            params={
                "page": page,
                "per_page": TABNEWS_PAGE_SIZE,
                "strategy": "new",
                "with_children": "true",
            },
        )

        if not items:
            break

        for item in items:
            title = item.get("title") or "[Comentario]"
            preview = select_text(item.get("body"), title)
            include, keywords_matched, _ = should_include_item(title, preview)
            if not include:
                continue

            detail = request_json(
                session,
                f"{TABNEWS_BASE_URL}/contents/{item['owner_username']}/{item['slug']}",
            )
            content = detail.get("body") or preview
            include_detail, detailed_keywords, _ = should_include_item(title, content)
            if not include_detail:
                continue

            records.append(
                build_record(
                    source="TabNews",
                    forum="Comments" if detail.get("parent_id") else "Posts",
                    author=str(detail.get("owner_username") or ""),
                    title=title,
                    content=content,
                    url=f"https://www.tabnews.com.br/{detail.get('owner_username')}/{detail.get('slug')}",
                    date=parse_iso_date(select_text(detail.get("published_at"), detail.get("created_at"))),
                    keywords_matched=detailed_keywords or keywords_matched,
                )
            )

    return deduplicate_records(records)


def extract_guj_topic_links(soup: BeautifulSoup) -> list[str]:
    links: list[str] = []
    for anchor in soup.select("a[href^='/t/']"):
        href = anchor.get("href", "")
        absolute_url = urljoin("https://www.guj.com.br", href)
        if absolute_url not in links:
            links.append(absolute_url)
        if len(links) >= GUJ_MAX_TOPIC_LINKS:
            break
    return links


def parse_guj_topic(session: requests.Session, url: str) -> dict[str, str] | None:
    soup = request_html(session, url)
    article = soup.select_one("article.topic-page")
    if article is None:
        return None

    title = select_text(article.select_one("header.topic-header h1").get_text(" ", strip=True) if article.select_one("header.topic-header h1") else "")
    category = select_text(
        article.select("nav.breadcrumb a")[-1].get_text(" ", strip=True)
        if len(article.select("nav.breadcrumb a")) >= 2
        else ""
    )
    original_post = article.select_one("div.post.post-original")
    if original_post is None:
        return None

    author = select_text(original_post.select_one(".post-author").get_text(" ", strip=True) if original_post.select_one(".post-author") else "")
    date_text = select_text(original_post.select_one("time").get_text(" ", strip=True) if original_post.select_one("time") else "")
    content = select_text(original_post.select_one(".post-body").get_text(" ", strip=True) if original_post.select_one(".post-body") else "")
    include, keywords_matched, _ = should_include_item(title, content)
    if not include:
        return None

    return build_record(
        source="GUJ",
        forum=category or "GUJ",
        author=author,
        title=title,
        content=content,
        url=url,
        date=parse_guj_date(date_text),
        keywords_matched=keywords_matched,
    )


def fetch_guj_leads() -> list[dict[str, str]]:
    session = build_session()
    records: list[dict[str, str]] = []

    response = session.get(GUJ_CATEGORY_URL, timeout=REQUEST_TIMEOUT)
    rate_limit_sleep()

    if response.ok:
        logging.info("Scanning GUJ category page: %s", GUJ_CATEGORY_URL)
        soup = BeautifulSoup(response.text, "html.parser")
    else:
        logging.warning(
            "GUJ category %s returned status %s. Falling back to %s",
            GUJ_CATEGORY_URL,
            response.status_code,
            GUJ_FALLBACK_URL,
        )
        soup = request_html(session, GUJ_FALLBACK_URL)

    topic_links = extract_guj_topic_links(soup)
    for topic_url in topic_links:
        try:
            record = parse_guj_topic(session, topic_url)
            if record:
                records.append(record)
        except Exception:
            logging.exception("Failed to parse GUJ topic: %s", topic_url)

    return deduplicate_records(records)


def main() -> None:
    configure_logging()

    collectors = {
        "Reddit": fetch_reddit_leads,
        "HackerNews": fetch_hackernews_leads,
        "TabNews": fetch_tabnews_leads,
        "GUJ": fetch_guj_leads,
    }

    records_by_source: dict[str, list[dict[str, str]]] = {source: [] for source in SOURCE_ORDER}

    for source_name, collector in collectors.items():
        try:
            logging.info("Starting source: %s", source_name)
            records = collector()
            records_by_source[source_name] = deduplicate_records(records)
            logging.info("Finished source: %s (%s leads)", source_name, len(records_by_source[source_name]))
        except Exception:
            logging.exception("Source failed: %s", source_name)
            records_by_source[source_name] = []

    export_to_excel(records_by_source, OUTPUT_FILE)
    total_records = sum(len(records_by_source.get(source, [])) for source in SOURCE_ORDER)
    logging.info("Exported %s total leads to %s", total_records, OUTPUT_FILE)


if __name__ == "__main__":
    main()
